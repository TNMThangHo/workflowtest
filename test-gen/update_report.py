"""
Update Test Report Script - Incremental report updates
Updates test execution results without regenerating entire report
"""

import pandas as pd
import json
import argparse
from datetime import datetime
from pathlib import Path
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

def parse_test_results(results_path: str, format_type: str) -> dict:
    """
    Parse test results from various formats
    Returns: dict mapping TestCase ID -> {status, actual_result, evidence}
    """
    print(f"📖 Parsing test results from: {results_path} (format: {format_type})")
    
    results_map = {}
    
    try:
        if format_type == 'json':
            with open(results_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Handle both list and dict formats
            if isinstance(data, list):
                for item in data:
                    tc_id = item.get('testcase_id') or item.get('TestCase ID')
                    results_map[tc_id] = {
                        'status': item.get('status') or item.get('Status'),
                        'actual_result': item.get('actual_result') or item.get('Actual Result'),
                        'evidence': item.get('evidence') or item.get('Evidence', '')
                    }
            else:
                # Assume dict with testcase_id as keys
                for tc_id, data in data.items():
                    results_map[tc_id] = {
                        'status': data.get('status'),
                        'actual_result': data.get('actual_result'),
                        'evidence': data.get('evidence', '')
                    }
        
        elif format_type in ['csv', 'excel']:
            if format_type == 'csv':
                df = pd.read_csv(results_path)
            else:
                df = pd.read_excel(results_path)
            
            for _, row in df.iterrows():
                tc_id = str(row.get('TestCase ID') or row.get('testcase_id'))
                results_map[tc_id] = {
                    'status': row.get('Status') or row.get('status'),
                    'actual_result': row.get('Actual Result') or row.get('actual_result'),
                    'evidence': row.get('Evidence') or row.get('evidence', '')
                }
        
        print(f"✅ Parsed {len(results_map)} test results")
        return results_map
        
    except Exception as e:
        print(f"❌ Error parsing results: {e}")
        sys.exit(1)

def validate_status(status: str) -> str:
    """Validate and normalize status values"""
    valid_statuses = ['Pass', 'Fail', 'Skip', 'Blocked']
    
    if pd.isna(status):
        return 'Not Executed'
    
    status = str(status).strip()
    
    # Case-insensitive matching
    for valid in valid_statuses:
        if status.lower() == valid.lower():
            return valid
    
    print(f"⚠️ Warning: Invalid status '{status}', using 'Not Executed'")
    return 'Not Executed'

def update_report(report_path: str, results_map: dict, output_dir: str):
    """
    Update test report with new execution results
    Preserves test case descriptions, only updates execution data
    """
    print("\n🔄 Updating test report with new results...")
    
    # Read existing report
    df_report = pd.read_excel(report_path)
    print(f"📖 Existing report: {len(df_report)} test cases")
    
    # Convert TestCase ID to string for matching
    df_report['TestCase ID'] = df_report['TestCase ID'].astype(str)
    
    # Track statistics
    updated_count = 0
    not_found_in_results = []
    not_found_in_report = []
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Update rows with new results
    for idx, row in df_report.iterrows():
        tc_id = row['TestCase ID']
        
        if tc_id in results_map:
            # Update execution data
            result = results_map[tc_id]
            df_report.at[idx, 'Status'] = validate_status(result['status'])
            df_report.at[idx, 'Actual Result'] = result['actual_result']
            
            if 'Evidence' in df_report.columns:
                df_report.at[idx, 'Evidence'] = result['evidence']
            
            if 'Execution Date' in df_report.columns:
                df_report.at[idx, 'Execution Date'] = today
            elif 'Last Execution Date' in df_report.columns:
                df_report.at[idx, 'Last Execution Date'] = today
            
            updated_count += 1
        else:
            not_found_in_results.append(tc_id)
    
    # Check for test results not in report
    for tc_id in results_map.keys():
        if tc_id not in df_report['TestCase ID'].values:
            not_found_in_report.append(tc_id)
    
    # Calculate statistics
    status_counts = df_report['Status'].value_counts().to_dict()
    total = len(df_report)
    
    pass_count = status_counts.get('Pass', 0)
    fail_count = status_counts.get('Fail', 0)
    skip_count = status_counts.get('Skip', 0)
    blocked_count = status_counts.get('Blocked', 0)
    not_executed = status_counts.get('Not Executed', 0)
    
    pass_rate = (pass_count / total * 100) if total > 0 else 0
    
    # Create backup
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = Path(output_dir) / f"backup_report_{timestamp}.xlsx"
    shutil.copy2(report_path, backup_path)
    print(f"💾 Backup created: {backup_path}")
    
    # Save updated report
    output_path = Path(output_dir) / f"updated_report_{timestamp}.xlsx"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write main test results
        df_report.to_excel(writer, index=False, sheet_name='Test Results')
        
        # Create summary sheet
        summary_data = {
            'Metric': [
                'Total Test Cases',
                'Updated in This Run',
                'Pass',
                'Fail',
                'Skip',
                'Blocked',
                'Not Executed',
                'Pass Rate (%)',
                'Execution Date'
            ],
            'Value': [
                total,
                updated_count,
                pass_count,
                fail_count,
                skip_count,
                blocked_count,
                not_executed,
                f"{pass_rate:.2f}%",
                today
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        
        # Apply formatting
        workbook = writer.book
        
        # Format Test Results sheet
        ws_results = writer.sheets['Test Results']
        
        try:
            # Copy formatting from original
            original_wb = load_workbook(report_path)
            original_ws = original_wb.active
            
            # Copy column widths
            for col_letter in original_ws.column_dimensions:
                if col_letter in ws_results.column_dimensions:
                    ws_results.column_dimensions[col_letter].width = \
                        original_ws.column_dimensions[col_letter].width
            
            # Copy header formatting
            for col_idx in range(1, len(df_report.columns) + 1):
                original_cell = original_ws.cell(row=1, column=col_idx)
                new_cell = ws_results.cell(row=1, column=col_idx)
                
                if original_cell.font:
                    new_cell.font = original_cell.font.copy()
                if original_cell.fill:
                    new_cell.fill = original_cell.fill.copy()
                if original_cell.alignment:
                    new_cell.alignment = original_cell.alignment.copy()
            
        except Exception as e:
            print(f"⚠️ Could not copy all formatting: {e}")
        
        # Format Summary sheet
        ws_summary = writer.sheets['Summary']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx in range(1, 3):
            cell = ws_summary.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    print(f"✅ Updated report saved: {output_path}")
    
    # Print summary
    print("\n" + "="*60)
    print("📊 TEST REPORT UPDATE SUMMARY")
    print("="*60)
    print(f"Total test cases: {total}")
    print(f"Updated: {updated_count} test cases")
    print(f"Not found in results: {len(not_found_in_results)}")
    if not_found_in_report:
        print(f"⚠️ TestCase IDs in results but not in report: {len(not_found_in_report)}")
    print(f"\n📈 Test Results:")
    print(f"   Pass: {pass_count} ({pass_count/total*100:.1f}%)")
    print(f"   Fail: {fail_count} ({fail_count/total*100:.1f}%)")
    print(f"   Skip: {skip_count} ({skip_count/total*100:.1f}%)")
    print(f"   Blocked: {blocked_count} ({blocked_count/total*100:.1f}%)")
    print(f"   Not Executed: {not_executed} ({not_executed/total*100:.1f}%)")
    print(f"\n🎯 Pass Rate: {pass_rate:.2f}%")
    print(f"🕐 Execution Date: {today}")
    print(f"\n📁 Files:")
    print(f"   - Updated: {output_path}")
    print(f"   - Backup: {backup_path}")
    
    if not_found_in_results:
        print(f"\n⚠️ TestCase IDs not found in results (unchanged):")
        for tc_id in not_found_in_results[:10]:  # Show first 10
            print(f"   - {tc_id}")
        if len(not_found_in_results) > 10:
            print(f"   ... and {len(not_found_in_results) - 10} more")
    
    if not_found_in_report:
        print(f"\n⚠️ TestCase IDs in results but not in report:")
        for tc_id in not_found_in_report[:10]:
            print(f"   - {tc_id}")
        if len(not_found_in_report) > 10:
            print(f"   ... and {len(not_found_in_report) - 10} more")
    
    print("="*60)
    
    return str(output_path)

def main():
    parser = argparse.ArgumentParser(description='Update test report with new execution results')
    parser.add_argument('--mode', choices=['parse', 'update'], required=True,
                       help='Mode: parse (parse results) or update (update report)')
    parser.add_argument('--results', help='Path to test results file (for parse mode)')
    parser.add_argument('--format', choices=['json', 'csv', 'excel'],
                       help='Format of results file (for parse mode)')
    parser.add_argument('--report', help='Path to existing test report (for update mode)')
    parser.add_argument('--parsed-results', help='Path to parsed results JSON (for update mode)')
    parser.add_argument('--output-dir', default='output', help='Output directory')
    
    args = parser.parse_args()
    
    if args.mode == 'parse':
        if not args.results or not args.format:
            print("❌ Error: --results and --format are required for parse mode")
            sys.exit(1)
        
        results_map = parse_test_results(args.results, args.format)
        
        # Save parsed results
        Path(args.output_dir).mkdir(parents=True, exist_ok=True)
        output_path = Path(args.output_dir) / 'parsed_results.json'
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results_map, f, ensure_ascii=False, indent=2)
        
        print(f"✅ Parsed results saved to: {output_path}")
        
    elif args.mode == 'update':
        if not args.report or not args.parsed_results:
            print("❌ Error: --report and --parsed-results are required for update mode")
            sys.exit(1)
        
        # Load parsed results
        with open(args.parsed_results, 'r', encoding='utf-8') as f:
            results_map = json.load(f)
        
        update_report(args.report, results_map, args.output_dir)

if __name__ == '__main__':
    main()
