"""
Quick script to convert existing raw_testcases.json to Excel format
This creates the initial Excel file needed for /update-tc workflow
"""
import pandas as pd
import json
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read existing testcases from JSON
with open('output/raw_testcases.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

test_cases = data['test_cases']

# Convert to DataFrame with Excel schema
excel_data = []
for tc in test_cases:
    steps_str = ' • '.join(tc.get('steps', [])) if isinstance(tc.get('steps'), list) else tc.get('steps', '')
    
    excel_data.append({
        'TestCase ID': tc.get('id', ''),
        'Module': tc.get('module', ''),
        'Type': tc.get('type', ''),
        'Priority': tc.get('priority', ''),
        'Title': tc.get('title', ''),
        'Pre-condition': tc.get('pre_condition', ''),
        'Test Steps': steps_str,
        'Expected Result': tc.get('expected_result', ''),
        'Status': 'Not Executed',
        'Actual Result': '',
        'Evidence': '',
        'Notes': tc.get('notes', ''),
        'Created Date': datetime.now().strftime('%Y-%m-%d')
    })

df = pd.DataFrame(excel_data)

# Save to Excel with formatting
output_path = Path('output/test_cases.xlsx')

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Test Cases')
    
    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Test Cases']
    
    # Apply header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    column_widths = {
        'A': 18,  # TestCase ID
        'B': 12,  # Module
        'C': 12,  # Type
        'D': 10,  # Priority
        'E': 40,  # Title
        'F': 20,  # Pre-condition
        'G': 50,  # Test Steps
        'H': 40,  # Expected Result
        'I': 15,  # Status
        'J': 40,  # Actual Result
        'K': 25,  # Evidence
        'L': 25,  # Notes
        'M': 12,  # Created Date
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Center align certain columns
    for row_idx in range(2, len(df) + 2):
        worksheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')  # ID
        worksheet.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')  # Type
        worksheet.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')  # Priority
        worksheet.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center')  # Status
        worksheet.cell(row=row_idx, column=13).alignment = Alignment(horizontal='center')  # Created Date
    
    # Enable text wrapping for long content
    for row_idx in range(1, len(df) + 2):
        for col_idx in [5, 6, 7, 8, 10, 11, 12]:  # Title, Pre, Steps, Expected, Actual, Evidence, Notes
            worksheet.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')

print(f"✅ Created Excel file: {output_path}")
print(f"   Total test cases: {len(df)}")
print(f"   Created Date: {datetime.now().strftime('%Y-%m-%d')}")
